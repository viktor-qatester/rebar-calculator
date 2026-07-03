"""Разбор загруженных файлов без ИИ — показать, что бот «видит» документ."""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

from bot.rebar_patterns import (
    count_rebar_mentions,
    find_rebar_snippets,
    find_rebar_snippets_by_page,
)

@dataclass
class FileInspection:
    ok: bool
    summary: str


def inspect_bytes(filename: str, data: bytes) -> FileInspection:
    suffix = Path(filename).suffix.lower()
    size_kb = len(data) / 1024

    if suffix == ".pdf":
        return _inspect_pdf(filename, data, size_kb)
    if suffix in {".jpg", ".jpeg", ".png", ".webp", ".bmp"}:
        return _inspect_image(filename, data, size_kb)
    if suffix in {".xlsx", ".xlsm", ".xltx"}:
        return _inspect_excel(filename, data, size_kb)
    if suffix in {".dwg", ".dxf"}:
        return _inspect_cad(filename, data, size_kb)

    return FileInspection(
        ok=False,
        summary=(
            f"📎 *{filename}*\n"
            f"Размер: {size_kb:.1f} КБ\n\n"
            "Формат пока не разбираю. Пришлите PDF, фото, Excel или DWG/DXF."
        ),
    )


def _extract_pdf_pages(data: bytes) -> tuple[list[str], int]:
    """pypdf, при неудаче — pymupdf (лучше читает КЖ)."""
    pages: list[str] = []
    try:
        from pypdf import PdfReader

        reader = PdfReader(BytesIO(data))
        for page in reader.pages:
            pages.append(page.extract_text() or "")
        return pages, len(reader.pages)
    except Exception:
        pass

    try:
        import fitz

        doc = fitz.open(stream=data, filetype="pdf")
        for page in doc:
            pages.append(page.get_text() or "")
        total = doc.page_count
        doc.close()
        return pages, total
    except ImportError:
        raise ImportError("Установите pypdf или pymupdf") from None


def _inspect_pdf(filename: str, data: bytes, size_kb: float) -> FileInspection:
    try:
        pages, page_count = _extract_pdf_pages(data)
    except ImportError:
        return FileInspection(ok=False, summary="Нет pypdf — установите зависимости бота.")
    except Exception as exc:
        return FileInspection(ok=False, summary=f"Не удалось открыть PDF: {exc}")

    text = "\n".join(pages)
    hits = find_rebar_snippets_by_page(pages) or find_rebar_snippets(text)
    mentions = count_rebar_mentions(text)

    lines = [
        f"📄 *{filename}*",
        "Тип: PDF",
        f"Размер: {size_kb:.1f} КБ · страниц: {page_count}",
        "Статус: *прочитан*",
    ]
    if text.strip():
        lines.append(f"Символов текста: {len(text.strip())}")
        lines.append(f"Совпадений (арматура, S400, сетки…): *{mentions}*")
        if hits:
            lines.append("\n*Нашёл в тексте:*")
            for h in hits:
                lines.append(f"• {h}")
        else:
            lines.append(
                "\nТекст есть, но маркировку арматуры не нашёл. "
                "Ищу: 12S400, S240, сетки, ø12, КЖ…"
            )
    else:
        lines.append("\n⚠️ Текст не извлечён — возможно, это скан. Пришлите PDF с текстовым слоем.")

    return FileInspection(ok=bool(text.strip()), summary="\n".join(lines))


def _inspect_image(filename: str, data: bytes, size_kb: float) -> FileInspection:
    w = h = None
    try:
        from PIL import Image

        img = Image.open(BytesIO(data))
        w, h = img.size
    except Exception:
        pass

    lines = [
        f"🖼 *{filename}*",
        "Тип: изображение",
        f"Размер файла: {size_kb:.1f} КБ",
        "Статус: *получено*",
    ]
    if w and h:
        lines.append(f"Разрешение: {w}×{h} px")

    lines.append(
        "\nOCR (чтение текста с фото) подключим позже. "
        "Сейчас бот подтверждает, что файл дошёл. Для разбора пришлите PDF с текстом."
    )
    return FileInspection(ok=True, summary="\n".join(lines))


def _inspect_excel(filename: str, data: bytes, size_kb: float) -> FileInspection:
    from bot.rebar_patterns import line_has_rebar

    try:
        from openpyxl import load_workbook
    except ImportError:
        return FileInspection(ok=False, summary="Нет модуля openpyxl — установите зависимости бота.")

    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    hits: list[str] = []
    for name in sheet_names[:5]:
        ws = wb[name]
        for row in ws.iter_rows(max_row=200, max_col=30, values_only=True):
            for cell in row:
                if cell is None:
                    continue
                s = str(cell)
                if line_has_rebar(s):
                    hits.append(f"{name}: {s[:100]}")
                    if len(hits) >= 5:
                        break
            if len(hits) >= 5:
                break
        if len(hits) >= 5:
            break
    wb.close()

    lines = [
        f"📊 *{filename}*",
        "Тип: Excel",
        f"Размер: {size_kb:.1f} КБ",
        f"Листы: {', '.join(sheet_names[:8])}" + ("…" if len(sheet_names) > 8 else ""),
        "Статус: *прочитан*",
    ]
    if hits:
        lines.append("\n*Нашёл:*")
        for h in hits:
            lines.append(f"• {h}")
    else:
        lines.append("\nЛисты открыты, но строк про арматуру не нашёл.")

    return FileInspection(ok=True, summary="\n".join(lines))


def _inspect_cad(filename: str, data: bytes, size_kb: float) -> FileInspection:
    from bot.rebar_patterns import line_has_rebar

    suffix = Path(filename).suffix.lower()
    if suffix == ".dxf":
        try:
            import ezdxf

            doc = ezdxf.read(BytesIO(data))
            texts: list[str] = []
            msp = doc.modelspace()
            for e in msp.query("TEXT MTEXT"):
                t = e.dxf.text if e.dxftype() == "TEXT" else e.text
                if t and line_has_rebar(str(t)):
                    texts.append(str(t)[:100])
                    if len(texts) >= 5:
                        break
            lines = [
                f"📐 *{filename}*",
                "Тип: DXF",
                f"Размер: {size_kb:.1f} КБ",
                "Статус: *прочитан*",
            ]
            if texts:
                lines.append("\n*Текст на чертеже:*")
                for t in texts:
                    lines.append(f"• {t}")
            else:
                lines.append("\nЧертёж открыт, подписей про арматуру не нашёл.")
            return FileInspection(ok=True, summary="\n".join(lines))
        except Exception as exc:
            return FileInspection(
                ok=False,
                summary=f"📐 *{filename}*\nНе удалось прочитать DXF: {exc}",
            )

    return FileInspection(
        ok=False,
        summary=(
            f"📐 *{filename}*\n"
            f"Размер: {size_kb:.1f} КБ\n\n"
            "DWG бинарный — для бесплатного разбора пришлите DXF или PDF выписку из проекта."
        ),
    )
